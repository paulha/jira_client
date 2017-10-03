from jira_class import Jira, get_query
from navigate import *
import sys
from os.path import expanduser, pathsep, dirname, realpath
from openpyxl import load_workbook
from jira_class import Jira
from jira.exceptions import JIRAError
import re
from utility_funcs.search import search_for_file

class XLS:
    SUMMARY = 1
    DESCRIPTION = 6

    def __init__(self, filename, log=None):
        self.workbook = load_workbook(filename)
        self.log = log

    def headings(self, sheet='Sheet1'):
        row = self.workbook[sheet][1]
        return row

    def read(self, sheet='Sheet1'):
        active_heading = self.headings(sheet=sheet)
        active_sheet = self.workbook[sheet]
        result = []
        for row in range(2, active_sheet.max_row+1):
            this_entry = {'ROW': row}
            this_row = active_sheet[row]
            for col in range(0, this_row.__len__()):
                this_entry[active_heading[col].value] = this_row[col].value
            result.append(this_entry)
        return result


class Supersede:
    def __init__(self, jira, parser=None, scenario=None, config=None, queries=None, search=None, log=None):
        self.jira = jira
        self.parser = parser
        self.scenario = scenario
        self.config = config
        self.queries = queries
        self.search = search
        self.log = log
        self.update = True if 'update' in scenario and scenario['update'] else False
        self.areq = None
        self.preq = None
        self.row = 0

    def check_access(self, items):
        self.row = items['ROW']
        self.areq = self.jira.jira_client.issue(items['AREQ'])
        if self.areq is None:
            self.log.logger.error("Unable to access issue %s at row %s", self.areq.key, self.row)

        self.preq = self.jira.jira_client.issue(items['PREQ'])
        if self.preq is None:
            self.log.logger.error("Unable to access issue %s at row %s", self.preq.key, self.row)

        return True if self.areq is not None and self.preq is not None else False

    def copy_areq_values_to_preq(self):
        """Copy areq to preq, overlaying data if present"""

        # -- TODO: shouldn't copy unless there's actually something to change...

        update_occured = 0

        # Utility function for copying *_on fields (see below)
        def _exists_on_update(update_list, field, source, target, tag=None):
            src = getattr(source.fields, field)
            src = {tag: getattr(x, tag) for x in (src if src is not None else [])}
            tgt = getattr(target.fields, field)
            tgt = {tag: getattr(x, tag) for x in (tgt if tgt is not None else [])}

            if 'exists_on' in self.scenario:
                _exists_on = self.scenario['exists_on']
                src[_exists_on] = _exists_on
            elif 'exists_only_on' in self.scenario:
                _exists_on = self.scenario['exists_only_on']
                src = {_exists_on: _exists_on}
            else:
                pass    # leave it alone!

            if tgt != src:      # tgt and src already can't be None (see above!)
                if tag is None:
                    # -- Good old fashioned assignment
                    update_list[field] = [x for x in src]
                else:
                    update_list[field] = [{tag: x} for key, x in src.items()]
                return True
            return False

        def _list_update(update_list, field, source, target, tag=None, remap=[]):

            src = getattr(source.fields, field) if getattr(source.fields, field) is not None else []
            src = [getattr(x, tag) for x in src] if tag is not None else src
            src = [(remap[x] if x in remap else x) for x in src]
            tgt = getattr(target.fields, field) if getattr(target.fields, field) is not None else []
            tgt = [getattr(x, tag) for x in tgt] if tag is not None else tgt
            if tgt.__str__() != src.__str__():
                if tag is None:
                    # -- Good old fashioned assignment
                    update_list[field] = src
                else:
                    update_list[field] = [{tag: x} for x in src] \
                                                   if src is not None else []
                return True
            return False

        def _field_update(update_list, field, source, target, tag=None):
            src = getattr(source.fields, field)
            tgt = getattr(target.fields, field)
            # FIXME: this has a problem because it's an object compare, not a value comparison.
            if tgt is None or tgt.__str__() != src.__str__():
                if tag is None:
                    # -- Good old fashioned assignment
                    update_list[field] = src
                else:
                    update_list[field] = {tag: getattr(src, tag)} if src is not None else {}
                return True
            return False

        exists_on = self.jira.get_field_name('Exists On')
        verified_on = self.jira.get_field_name('Verified On')
        failed_on = self.jira.get_field_name('Failed On')
        blocked_on = self.jira.get_field_name('Blocked On')
        tested_on = self.jira.get_field_name('Tested On')
        validation_lead = self.jira.get_field_name('Validation Lead')
        classification = self.jira.get_field_name('Classification')

        val_lead = getattr(self.areq.fields, validation_lead)

        update_assignee_dict = {
            # 'assignee': {'name': self.areq.fields.assignee.name},
        }
        _field_update(update_assignee_dict, 'assignee', self.areq, self.preq, 'name')

        update_lead_dict = {
            # validation_lead: {'name': val_lead.name if val_lead is not None else ""}
        }
        _field_update(update_lead_dict, validation_lead, self.areq, self.preq, 'name')

        # -- Having created the issue, now other fields of the E-Feature can be updated:
        update_fields = {
            # 'priority': {'name': self.areq.fields.priority.name if self.areq is not None else 'P1-Stopper'},
        }
        _field_update(update_fields, 'priority', self.areq, self.preq, 'name')
        _list_update(update_fields, 'components', self.areq, self.preq, 'name')
        _list_update(update_fields, 'labels', self.areq, self.preq)
        _list_update(update_fields, classification, self.areq, self.preq, 'value',
                     remap={'Functional': 'Functional Use Case'})
        _list_update(update_fields, verified_on, self.areq, self.preq, 'value')
        _list_update(update_fields, failed_on, self.areq, self.preq, 'value')
        _list_update(update_fields, blocked_on, self.areq, self.preq, 'value')
        _list_update(update_fields, tested_on, self.areq, self.preq, 'value')

        # FIXME: What, exactly, is this code doing? (I know I wrote it, but...)
        #
        # 1. Get current values from both items
        # 2. If source doesn't have the target in
        #    it yet, add it (or set it).
        # 3. Normalize both source and target to
        #    a standardized list form
        # 4. Compare the two. If they are different
        #    then target must be updated with source
        # 5. Set the target list.
        # -----------------------------------------
        # if 'exists_on' in self.scenario:
        #     exists_list = self.scenario['exists_on']
        # elif 'exists_only_on' in self.scenario:
        #     exists_list = self.scenario['exists_only_on']
        # else:
        #     exists_list = None
        #
        # if exists_list:
        #     exists_on_list = [{'value': exists_list}]
        #     if 'exists_on' in self.scenario:
        #         exists_on_list.__add__([{'value': x.value} for x in getattr(self.areq.fields, exists_on)])
        #
        #     update_fields[exists_on] = exists_on_list
        _exists_on_update(update_fields, exists_on, self.areq, self.preq, 'value')


        # -- Create the e-feature and update the stuff you can't set directly
        updated = False
        if self.update:
            if update_fields:
                self.log.logger.info("Updating Fields from %s to %s" % (self.areq.key, self.preq.key))
                self.preq.update(notify=False, fields=update_fields)
                updated = True
            if update_assignee_dict:
                try:
                    self.log.logger.info("Updating Assignee from %s to %s" % (self.areq.key, self.preq.key))
                    self.preq.update(notify=False, fields=update_assignee_dict)
                    updated = True
                except JIRAError as e:
                    self.log.logger.error("Jira error %s", e)
            if update_lead_dict:
                try:
                    self.log.logger.info("Updating Lead from %s to %s" % (self.areq.key, self.preq.key))
                    self.preq.update(notify=False, fields=update_lead_dict)
                    updated = True
                except JIRAError as e:
                    self.log.logger.error("Jira error %s", e)

            # -- Add a comment noting the creation of this feature.
            if updated:
                self.log.logger.info("Leaving comments on %s and %s" % (self.areq.key, self.preq.key))
                self.jira.jira_client.add_comment(self.areq,
                                                  """This item was superseded by {command}.
        
                                                  Superseded item is %s. 
                                                  Replacement item is %s
        
                                                  %s""".format_map(self.scenario)
                                                  % (self.areq.key, self.preq.key,
                                                     self.scenario['comment'] if self.scenario['comment']
                                                                                 is not None else ""))
                self.jira.jira_client.add_comment(self.preq,
                                                  """This item supersedes %s by {command}.
    
                                                  This item is %s
    
                                                  %s""".format_map(self.scenario)
                                                  % (self.areq.key, self.preq.key,
                                                     self.scenario['comment'] if self.scenario['comment']
                                                                                 is not None else ""))

                self.log.logger.info("Updated PREQ %s from %s: ", self.preq.key, self.areq.key)

        return updated

    def create_areq_duplicate_of_preq_link(self):
        updated = False

        # -- Check to see if there's a link between this areq and preq
        has_duplicates_link = [link.outwardIssue
                               for link in self.areq.fields.issuelinks
                               if hasattr(link, "outwardIssue")
                                  and link.type.name == "Duplicate"
                                  and link.outwardIssue.key == self.preq.key
                               ]

        if not has_duplicates_link:
            if self.update:
                self.log.logger.info("Creating link from %s to %s" % (self.areq.key, self.preq.key))
                self.jira.create_issue_link("Duplicate", self.areq, self.preq)
                self.log.logger.info("Created 'Duplicates' link: %s --> %s", self.areq.key, self.preq.key)
                updated = True
            else:
                self.log.logger.warning("Link from %s --> %s is MISSING", self.areq.key, self.preq.key)

        return updated

    def set_areq_to_Rejected(self):

        updated = False
        if self.update:
            if self.areq.fields.status.name != 'Rejected':
                self.log.logger.info("Setting %s to Rejected" % (self.areq.key))
                target_state = {
                    'Feature': Feature_Rejected,
                    'E-Feature': E_Feature_Reject,
                    'UCIS': UCIS_Rejected,
                }[self.areq.fields.issuetype.name]
                StateMachine.transition_to_state(self.jira, self.areq, target_state, self.log)
                updated = True
            else:
                self.log.logger.info("%s is already Rejected" % (self.areq.key))

        return updated

    def supersede(self, items):
        updated = False
        if not self.check_access(items):
            self.log.logger.error("Abandoning row...")
            return False
        updated = self.copy_areq_values_to_preq() or updated
        updated = self.create_areq_duplicate_of_preq_link() or updated
        # todo: You could also set the state of the target...
        updated = self.set_areq_to_Rejected() or updated

        return updated




def areq_superceded_by_preq(parser, scenario, config, queries, search, log=None):
    """Copy values from areq to preq, Link areq to preq as 'Duplicates', then Reject areq, leave comment on areq and preq.

    Also:

    * Make sure areq and preq don't already have a 'Duplicates' link!
    * If areq status is already 'Reject', no need to re-set it.
    * If corresponding values are already set, no need to re-set them.
    * Should target values be overwritten if they are set (kinda think not...)
    * If areq status fails to be 'Reject', flag clearly!
    """

    update = scenario['update']
    log.logger.info("Update is %s", update)
    log.logger.info("Marking AREQ items as superseded by PREQ item.")
    log.logger.info("=================================================================")

    updates = 0

    XLS_FILE = realpath(dirname(realpath(sys.argv[0])) + '/../' + scenario['xls_input'])
    log.logger.info("input file: %s", XLS_FILE)
    xls = XLS(XLS_FILE)
    items = xls.read()

    # -- Get and format it:
    jira = Jira(scenario['name'], search, log=log.logger)

    issue = jira.jira_client.issue("AREQ-23223")

    supersede = Supersede(jira, parser, scenario, config, queries, search, log=log)
    updates = 0
    for item in items:
        log.logger.info("Processing %s, %s and %s", item['ROW'], item['AREQ'], item['PREQ'])
        changed = supersede.supersede(item)

        updates = updates + 1 if changed else updates

        if scenario['createmax'] and updates >= scenario['createmax']:
            break

    log.logger.info("-----------------------------------------------------------------")
    log.logger.info("%s items were updated ", updates)
    log.logger.info("")

    return


