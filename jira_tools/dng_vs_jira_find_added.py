import sys
from os.path import expanduser, pathsep, dirname, realpath
from openpyxl import load_workbook
from jira_class import Jira, get_query, remove_version_and_platform, strip_non_ascii
import re
from utility_funcs.search import search_for_file

class Utility:

    def _add_to_dict(self, dictionary, key, value, key_name=""):
        if key not in dictionary:
            dictionary[key] = value
        else:
            if key_name not in dictionary[key]['duplicates']:
                dictionary[key]['duplicates'][key_name] = []
            dictionary[key]['duplicates'][key_name].append(value)
            raise KeyError("Duplicate key: '%s': %s" % (key, dictionary[key]))

    def _gather_duplications(self, dictionary, key_name=""):
        return [entry for key, entry in dictionary.items()
                if key_name in entry['duplicates'] and len(entry['duplicates'][key_name]) > 0]

    def add_match(self, this_item, matching_item, key_name):
        matched_by = this_item['matchedby']
        if key_name not in matched_by:
            matched_by[key_name] = []
        matched_by[key_name].append(matching_item)


class DNG_File (Utility):
    SUMMARY = 1
    DESCRIPTION = 6

    def __init__(self, filename=None, log=None):
        self.workbook = None
        self.entries = []
        self.dng_by_summary = {}
        self.dng_by_description = {}

        if isinstance(filename, str):
            self.read(filename, log=log)

    def _empty_entry(self):
        return {
            'lineno':       None,
            'summary':      None,
            'description':  None,
            'row':          None,
            'matchedby':    {},
            'duplicates':   {},
        }

    def gather_summary_duplications(self):
        return self._gather_duplications(self.dng_by_summary, "summary")

    def gather_description_duplications(self):
        return self._gather_duplications(self.dng_by_description, "description")

    def read(self, filename, log=None):
        self.workbook = load_workbook(filename)
        sheet1 = self.workbook['Sheet1']
        for row in range(2, sheet1.max_row):
            this_row = sheet1[row]
            entry = self._empty_entry()
            entry['lineno'] = row
            entry['summary'] = this_row[self.SUMMARY].value
            entry['description'] = this_row[self.DESCRIPTION].value
            entry['row'] = this_row
            self.entries.append(entry)
            try:
                self._add_to_dict(self.dng_by_summary, entry['summary'], entry, "summary")
            except KeyError as e:
                pass
                # log.logger.warning("%s", e)

            try:
                self._add_to_dict(self.dng_by_description, entry['description'], entry, "description")
            except KeyError as e:
                pass
                # log.logger.warning("%s", e)

    def get_entries(self):
        return self.entries

class Jira_Project (Jira, Utility):

    def __init__(self, server_alias, config_path, log=None):
        self.items = None
        self.items_by_key = {}
        self.items_by_summary = {}
        self.items_by_description = {}

        super().__init__(server_alias, config_path, log)

    def _empty_item(self):
        return {
            'key':          None,
            'summary':      None,
            'description':  None,
            'item':         None,
            'matchedby':    {},
            'duplicates':   {},
        }

    def gather_key_duplications(self):
        return self._gather_duplications(self.items_by_key, "key")

    def gather_summary_duplications(self):
        return self._gather_duplications(self.items_by_summary, "summary")

    def gather_description_duplications(self):
        return self._gather_duplications(self.items_by_description, "description")

    def read(self, query, log=None):
        self.items = []
        for item in self.do_query(query):
            new_item = self._empty_item()
            new_item['key'] = item.key
            new_item['summary'] = strip_non_ascii(remove_version_and_platform(item.fields.summary))
            new_item['description'] = strip_non_ascii(item.fields.description)
            new_item['item'] = item
            self.items.append(new_item)
            try:
                self._add_to_dict(self.items_by_key, new_item['key'], new_item, "key")
            except KeyError as e:
                pass
                # log.logger.warning("%s", e)

            try:
                self._add_to_dict(self.items_by_summary, new_item['summary'], new_item, "summary")
            except KeyError as e:
                pass
                # log.logger.warning("%s", e)

            try:
                self._add_to_dict(self.items_by_description, new_item['description'], new_item, "description")
            except KeyError as e:
                pass
                # log.logger.warning("%s", e)


def display_duplications(title, id_key_field, index_field_name, dictionary, log=None):
    if len(dictionary) > 0:
        fmt = "%-15s %-10s %-80s"
        log.logger.warning("")
        log.logger.warning(title.format_map({'count': len(dictionary)}))
        log.logger.warning("")
        log.logger.warning(fmt, id_key_field.center(15), "Duplicates".center(10), index_field_name.center(40))
        log.logger.warning(fmt, "=" * 15, "=" * 10, "=" * 80)
        for item in dictionary:
            # dups = [item[id_key_field]]
            dups = str(item[id_key_field])
            for k, d in item['duplicates'].items():
                for entry in d:
                    # dups.append(entry[id_key_field])
                    dups += " " + str(entry[id_key_field])
            log.logger.warning(fmt, dups, index_field_name, escape("'" + item[index_field_name] + "'"))
        log.logger.warning("")
        log.logger.warning("-" * 10)

    else:
        log.logger.info("")
        log.logger.info(title.format_map({'count': 0}))
        log.logger.info("")
        log.logger.info("=====================================================================================")
        log.logger.info("")
        log.logger.info("  -- No Duplications Found --")
        log.logger.info("")

def escape(text):
    return text.translate(str.maketrans({"\n":"\\n", "\r": "\\r"}))

def look_for_matches(dng, jira, log=None):
    for dng_item in dng.get_entries():
        match_found = False
        if dng_item['summary'] in jira.items_by_summary:
            # found a match!
            jira_item = jira.items_by_summary[dng_item['summary']]
            dng.add_match(dng_item, jira_item, "summary")
            jira.add_match(jira_item, dng_item, "summary")
            log.logger.debug("Found a summary match %s to DNG %s", jira_item['key'], dng_item['row'][0].value)
            match_found = True

        if dng_item['description'] in jira.items_by_description:
            # found a match!
            jira_item = jira.items_by_description[dng_item['description']]
            dng.add_match(dng_item, jira_item, "description")
            jira.add_match(jira_item, dng_item, "description")
            log.logger.debug("Found a description match %s to DNG %s", jira_item['key'], dng_item['row'][0].value)
            match_found = True

        if not match_found:
            log.logger.debug("NO MATCH FOUND -- DNG item: %d: %s", dng_item['row'][0].value, dng_item['row'][1].value)


def display_jira_matches(title, item_list, log=None):
    matches = [item for item in item_list if len(item['matchedby']) > 0]
    logger = log.logger.info if len(matches) > 0 else log.logger.warn
    fmt = "%-10s %-11s %-10s %-100s"
    logger("")
    logger(title.format_map({'count': len(matches)}))
    logger("")
    logger(fmt, "Key".center(10), "Match Type".center(11), "Matches".center(10), "Summary".center(40))
    logger(fmt, "=" * 10, "=" * 11, "=" * 10, "=" * 100)
    for item in matches:
        item_key = item['key']
        for match_type, match_list in item['matchedby'].items():
            matched_list = ""
            for matching_item in match_list:
                matched_list += str(matching_item['lineno'])+" "
                logger(fmt, item_key, match_type, matched_list, escape("'"+item['summary']+"'")[:100])
    logger("")
    logger("-" * 10)


def display_dng_match_failures(title: str, item_list: list, log: logging.Logger=None) -> None:
    non_matches = [item for item in item_list if len(item['matchedby']) == 0]
    logger = log.logger.warn if len(non_matches) > 0 else log.logger.info
    fmt = "%-6s %-100s"
    logger("")
    logger(title.format_map({'count': len(non_matches)}))
    logger("")
    logger(fmt, "Row".center(6), "Summary".center(20))
    logger(fmt, "=" * 6, "=" * 100)
    for item in non_matches:
        item_row = item['lineno']
        item_description = item['description']
        logger(fmt, item_row, escape("'"+item['summary']+"'")[:100])
        logger(fmt, "", escape("'" + str(item['description']) + "'")[:100])
    logger("")
    logger("-" * 10)


def find_added_dng_vs_jira(parser, scenario, config, queries, search, log=None):
    """Scan DNG input, label Jira entries and Note DNG items not in Jira"""

    preq_source_query = get_query('preq_source_query', queries, find_added_dng_vs_jira.__name__, params=scenario, log=log)

    verify = scenario['verify']
    update = scenario['update']
    log.logger.info("Verify is %s and Update is %s", verify, update)
    log.logger.info("=================================================================")

    XLS_FILE = realpath(dirname(realpath(sys.argv[0])) + '/../KSL-I Android.xlsx')
    dng = DNG_File(XLS_FILE, log=log)
    log.logger.info("Read %d DNG entries", len(dng.entries)-1)  # Account for column heading...
    jira = Jira_Project(scenario['name'], search, log=log.logger)
    jira.read(preq_source_query)
    log.logger.info("Read %d jira entries", len(jira.items))

    summary_duplications = dng.gather_summary_duplications()
    display_duplications("{count} Duplications of summary found in the input worksheet:",
                         "lineno", "summary", summary_duplications, log=log)
    description_duplications = dng.gather_description_duplications()
    display_duplications("{count} Duplications of description found in the input worksheet:",
                         "lineno", "description", description_duplications, log=log)

    jira_key_duplications = jira.gather_key_duplications()
    display_duplications("{count} Duplications of Jira Keys found in the input project:",
                         "key", "key", jira_key_duplications, log=log)

    jira_summary_duplications = jira.gather_summary_duplications()
    display_duplications("{count} Duplications of Jira Summaries found in the input project:",
                         "key", "summary", jira_summary_duplications, log=log)

    jira_description_duplications = jira.gather_description_duplications()
    display_duplications("{count} Duplications of Jira Descriptions found in the input project:",
                         "key", "description", jira_description_duplications, log=log)

    look_for_matches(dng, jira, log)

    display_jira_matches("These {count} jira items matched DNG entries", jira.items_by_key.values(), log=log)
    display_dng_match_failures("These {count} DNG items had no matcing Jira PREQ", dng.entries, log=log)

    source_preq_scanned = 0
    source_areq_scanned = 0
    ucis_created = 0
    e_features_created = 0
    warnings_issued = 0
    verify_failures = 0
    update_failures = 0
    processing_errors = 0

    update_count = 0


    # log.logger.info("-----------------------------------------------------------------")
    # log.logger.info("%s UCIS source entries were considered. ", source_preq_scanned)
    # log.logger.info("%s target UCIS entries were created. ", ucis_created)
    # log.logger.info("%s E-Feature source entries were considered. ", source_areq_scanned)
    # log.logger.info("%s target E-Features entries were created. ", e_features_created)
    # log.logger.info("%s warnings were issued. ", warnings_issued)
    # log.logger.info("")
    #
    # log.logger.info("%s processing error(s). ", processing_errors)
